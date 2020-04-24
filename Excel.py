# Import module
import openpyxl



# Open excel
wb = openpyxl.load_workbook('BD2.xlsx')

# Select table
sh = wb['Planilha1']

# Min and max row/column
coluns = sh.max_column
rows = sh.max_row

# Search word
word = input('Search: ')

# Load all table
for i in range(1, rows + 1):
    for j in range(1, coluns + 1):
        c = sh.cell(i, j)
        # Find the search
        if str(c.value).lower().find(word.lower()) != -1:
            for x in range(1, coluns + 1):
                    bd1 = sh.cell(i, x).value


